# -- coding: utf-8 --
__author__ = "ruijing.li"
__version__ = "0.1.0"

import pandas as pd
import pyreadstat
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Font
from openpyxl.styles import PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

def set_worksheet_format(ws):
    ws.row_dimensions[1].height = 50
    for col in ws.columns:
        index = list(ws.columns).index(col)
        letter = get_column_letter(index + 1)
        ws.column_dimensions[letter].width = 14
        ws[f"{letter}1"].alignment = Alignment(wrapText=True)
    ws.auto_filter.ref = ws.dimensions


#写一个把df输出到excel的方法，需要在输出的时候用空字符串替换“Null”
def write_excel(df, path, sheet_name):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    for r in dataframe_to_rows(df, index=True, header=True):
        ws.append(r)
    for row in ws.iter_rows(min_row=1, max_row=1):
        for cell in row:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="DDDDDD")
    set_worksheet_format(ws)
    wb.save(path)
    return wb
